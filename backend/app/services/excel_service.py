"""
Excel Export Service — ConstructVision AI
Generates a styled multi-sheet Excel workbook.
Sheet 1: Summary Dashboard
Sheet 2: Full BOQ
Sheet 3: Category Totals
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint


# ── Colour palette ─────────────────────────────────────────────────────────
C_BRAND_ORANGE = "FFFF7510"
C_DARK_BG      = "FF1A1714"
C_SLATE_LIGHT  = "FFF6F5F4"
C_SLATE_MID    = "FFD2CEC9"
C_SLATE_DARK   = "FF47423A"
C_WHITE        = "FFFFFFFF"
C_MUTED        = "FF7A7269"
C_BLUE         = "FF3B82F6"
C_PURPLE       = "FF8B5CF6"
C_AMBER        = "FFF59E0B"
C_CYAN         = "FF06B6D4"
C_GREEN        = "FF10B981"
C_RED          = "FFEF4444"

CATEGORY_COLOURS = {
    "civil":        C_BLUE,
    "civil_work":   C_BLUE,
    "finishing":    C_PURPLE,
    "electrical":   C_AMBER,
    "plumbing":     C_CYAN,
    "external":     C_GREEN,
    "external_work":C_GREEN,
}
CATEGORY_LABELS = {
    "civil":"Civil Work","civil_work":"Civil Work","finishing":"Finishing Work",
    "electrical":"Electrical Work","plumbing":"Plumbing Work",
    "external":"External Work","external_work":"External Work",
}
CAT_ORDER = ["civil","civil_work","plumbing","electrical","finishing","external","external_work"]


def _fill(hex_colour: str):
    return PatternFill("solid", fgColor=hex_colour)

def _font(bold=False, colour=C_DARK_BG, size=10, italic=False):
    return Font(bold=bold, color=colour, size=size, name="Calibri", italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style="thin"):
    s = Side(style=style, color="FFD2CEC9")
    return Border(left=s, right=s, top=s, bottom=s)

def _fmt_inr(val) -> str:
    if val is None: return "—"
    n = float(val)
    if n >= 10_000_000: return f"₹{n/10_000_000:.2f} Cr"
    if n >= 100_000:    return f"₹{n/100_000:.2f} L"
    return f"₹{n:,.0f}"

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def generate_boq_excel(project, estimation) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    _build_summary_sheet(wb, project, estimation)
    _build_boq_sheet(wb, project, estimation)
    _build_category_sheet(wb, estimation)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _build_summary_sheet(wb: Workbook, project, estimation):
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_view.showGridLines = False

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = "ConstructVision AI — Project Cost Summary"
    ws["A1"].fill = _fill(C_DARK_BG)
    ws["A1"].font = Font(bold=True, color=C_BRAND_ORANGE[2:], size=14, name="Calibri")
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = project.name
    ws["A2"].fill = _fill(C_DARK_BG)
    ws["A2"].font = Font(bold=True, color="FFD2CEC9", size=11, name="Calibri")
    ws["A2"].alignment = _align("center")
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 10

    # ── Project details ────────────────────────────────────────────────────────
    def write_detail(row, label, value):
        ws.cell(row=row, column=1, value=label).font = _font(bold=True, colour=C_MUTED, size=9)
        ws.cell(row=row, column=1).fill = _fill(C_SLATE_LIGHT)
        ws.cell(row=row, column=1).alignment = _align()
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = _font(bold=True, size=10)
        cell.fill = _fill(C_WHITE)
        cell.alignment = _align()
        for c in range(1, 3):
            ws.cell(row=row, column=c).border = _border()

    details = [
        ("Project Name", project.name),
        ("Project Type", project.project_type.capitalize()),
        ("Location", ", ".join(filter(None,[project.city,project.state])) or "—"),
        ("Built-up Area", f"{float(project.total_area_sqft):,.0f} sq.ft" if project.total_area_sqft else "—"),
        ("No. of Floors", str(project.num_floors)),
        ("Finish Quality", project.finish_quality.capitalize()),
        ("Date Prepared", datetime.now().strftime("%d %b %Y")),
        ("AI Confidence", (estimation.ai_confidence or "medium").capitalize()),
    ]
    for i, (lbl, val) in enumerate(details):
        write_detail(4 + i, lbl, val)

    ws.row_dimensions[12].height = 10

    # ── Cost breakdown table ──────────────────────────────────────────────────
    ws.merge_cells("A13:F13")
    ws["A13"] = "COST BREAKDOWN"
    ws["A13"].fill = _fill(C_DARK_BG)
    ws["A13"].font = Font(bold=True, color=C_WHITE[2:], size=10, name="Calibri")
    ws["A13"].alignment = _align("center")
    ws.row_dimensions[13].height = 20

    headers = ["Category", "Amount (₹)", "% of Total"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=14, column=ci, value=h)
        cell.fill = _fill(C_SLATE_MID)
        cell.font = _font(bold=True, size=9)
        cell.alignment = _align("center")
        cell.border = _border()

    breakdown_map = {
        "civil_work_cost": "Civil Work",
        "finishing_cost":  "Finishing Work",
        "electrical_cost": "Electrical Work",
        "plumbing_cost":   "Plumbing Work",
        "contingency_cost":"Contingency (5%)",
    }
    total = float(estimation.total_cost or 1)

    row = 15
    for attr, label in breakdown_map.items():
        val = float(getattr(estimation, attr, None) or 0)
        pct = (val / total * 100) if total > 0 else 0
        ws.cell(row=row, column=1, value=label).font = _font(size=9)
        ws.cell(row=row, column=1).border = _border()
        ws.cell(row=row, column=1).fill = _fill(C_SLATE_LIGHT if row%2==0 else C_WHITE)

        amt_cell = ws.cell(row=row, column=2, value=val)
        amt_cell.number_format = '₹#,##0.00'
        amt_cell.font = _font(size=9)
        amt_cell.alignment = _align("right")
        amt_cell.border = _border()
        amt_cell.fill = _fill(C_SLATE_LIGHT if row%2==0 else C_WHITE)

        pct_cell = ws.cell(row=row, column=3, value=round(pct, 1))
        pct_cell.number_format = '0.0"%"'
        pct_cell.font = _font(size=9)
        pct_cell.alignment = _align("center")
        pct_cell.border = _border()
        pct_cell.fill = _fill(C_SLATE_LIGHT if row%2==0 else C_WHITE)
        row += 1

    # Grand total row
    ws.cell(row=row, column=1, value="TOTAL PROJECT COST").font = Font(bold=True, color=C_WHITE[2:], size=10, name="Calibri")
    ws.cell(row=row, column=1).fill = _fill(C_DARK_BG)
    ws.cell(row=row, column=1).border = _border()

    tot_cell = ws.cell(row=row, column=2, value=total)
    tot_cell.number_format = '₹#,##0.00'
    tot_cell.font = Font(bold=True, color=C_BRAND_ORANGE[2:], size=11, name="Calibri")
    tot_cell.alignment = _align("right")
    tot_cell.fill = _fill(C_DARK_BG)
    tot_cell.border = _border()

    ws.cell(row=row, column=3, value="100%").font = Font(bold=True, color=C_WHITE[2:], size=9, name="Calibri")
    ws.cell(row=row, column=3).fill = _fill(C_DARK_BG)
    ws.cell(row=row, column=3).alignment = _align("center")
    ws.cell(row=row, column=3).border = _border()

    # Cost per sqft
    if estimation.cost_per_sqft:
        ws.row_dimensions[row+2].height = 10
        ws.cell(row=row+3, column=1, value="Cost per sq.ft").font = _font(bold=True, colour=C_MUTED, size=9)
        cpf_cell = ws.cell(row=row+3, column=2, value=float(estimation.cost_per_sqft))
        cpf_cell.number_format = '₹#,##0.00'
        cpf_cell.font = _font(bold=True, size=11)
        cpf_cell.alignment = _align("right")

    # ── Column widths ──────────────────────────────────────────────────────────
    _set_col_width(ws, "A", 28)
    _set_col_width(ws, "B", 20)
    _set_col_width(ws, "C", 14)
    for col in ["D","E","F"]:
        _set_col_width(ws, col, 12)


def _build_boq_sheet(wb: Workbook, project, estimation):
    ws = wb.create_sheet("📋 Bill of Quantities")
    ws.sheet_view.showGridLines = False

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"BILL OF QUANTITIES — {project.name.upper()}"
    ws["A1"].fill = _fill(C_DARK_BG)
    ws["A1"].font = Font(bold=True, color=C_BRAND_ORANGE[2:], size=12, name="Calibri")
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    loc = ", ".join(filter(None,[project.city, project.state])) or "India"
    ws["A2"] = f"{project.project_type.capitalize()}  ·  {loc}  ·  {float(project.total_area_sqft or 0):,.0f} sq.ft  ·  {project.finish_quality.capitalize()} Finish"
    ws["A2"].fill = _fill(C_SLATE_DARK)
    ws["A2"].font = Font(color=C_SLATE_MID[2:], size=9, name="Calibri")
    ws["A2"].alignment = _align("center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = ["Item Code","Description of Work","Unit","Quantity","Rate (₹)","Amount (₹)","Edited"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.fill = _fill(C_DARK_BG)
        cell.font = Font(bold=True, color=C_WHITE[2:], size=9, name="Calibri")
        cell.alignment = _align("center")
        cell.border = _border()
    ws.row_dimensions[4].height = 20

    # Freeze panes so header stays visible
    ws.freeze_panes = "A5"

    # ── Group items by category ────────────────────────────────────────────────
    grouped: dict[str, list] = {}
    for item in estimation.boq_items:
        c = item.category.lower()
        grouped.setdefault(c, []).append(item)

    cats = sorted(grouped.keys(), key=lambda x: CAT_ORDER.index(x) if x in CAT_ORDER else 99)

    row = 5
    grand_total = 0.0

    for cat in cats:
        items = grouped[cat]
        cat_total  = sum(float(i.amount or 0) for i in items)
        grand_total += cat_total
        cat_colour  = CATEGORY_COLOURS.get(cat, C_SLATE_DARK)
        cat_label   = CATEGORY_LABELS.get(cat, cat.title())

        # Category header
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1, value=f"  {cat_label.upper()}").font = Font(
            bold=True, color=C_WHITE[2:], size=9, name="Calibri"
        )
        ws.cell(row=row, column=1).fill = _fill(cat_colour)
        ws.cell(row=row, column=1).alignment = _align()

        amt_cell = ws.cell(row=row, column=6, value=cat_total)
        amt_cell.number_format = '₹#,##0.00'
        amt_cell.font = Font(bold=True, color=C_WHITE[2:], size=9, name="Calibri")
        amt_cell.fill = _fill(cat_colour)
        amt_cell.alignment = _align("right")
        ws.row_dimensions[row].height = 18
        row += 1

        for i, item in enumerate(items):
            bg = C_WHITE if i % 2 == 0 else C_SLATE_LIGHT

            ws.cell(row=row, column=1, value=item.item_code or "—").font = Font(color=C_MUTED[2:], size=8, name="Calibri Monospace")
            ws.cell(row=row, column=1).fill = _fill(bg)
            ws.cell(row=row, column=1).alignment = _align("center")
            ws.cell(row=row, column=1).border = _border()

            desc_cell = ws.cell(row=row, column=2, value=item.description)
            desc_cell.font = _font(size=9)
            desc_cell.fill = _fill(bg)
            desc_cell.alignment = _align(wrap=True)
            desc_cell.border = _border()

            ws.cell(row=row, column=3, value=str(item.unit).upper()).font = Font(color=C_MUTED[2:], size=8, name="Calibri")
            ws.cell(row=row, column=3).fill = _fill(bg)
            ws.cell(row=row, column=3).alignment = _align("center")
            ws.cell(row=row, column=3).border = _border()

            qty_cell = ws.cell(row=row, column=4, value=float(item.quantity or 0))
            qty_cell.number_format = '#,##0.000'
            qty_cell.font = _font(size=9)
            qty_cell.fill = _fill(bg)
            qty_cell.alignment = _align("right")
            qty_cell.border = _border()

            rate_cell = ws.cell(row=row, column=5, value=float(item.rate or 0))
            rate_cell.number_format = '₹#,##0.00'
            rate_cell.font = _font(size=9)
            rate_cell.fill = _fill(bg)
            rate_cell.alignment = _align("right")
            rate_cell.border = _border()

            amt_cell = ws.cell(row=row, column=6, value=float(item.amount or 0))
            amt_cell.number_format = '₹#,##0.00'
            amt_cell.font = _font(bold=True, size=9)
            amt_cell.fill = _fill(bg)
            amt_cell.alignment = _align("right")
            amt_cell.border = _border()

            edited_cell = ws.cell(row=row, column=7, value="✓" if item.is_user_edited else "")
            edited_cell.font = Font(color="FF10B981", size=9, name="Calibri")
            edited_cell.fill = _fill(bg)
            edited_cell.alignment = _align("center")
            edited_cell.border = _border()

            ws.row_dimensions[row].height = 16
            row += 1

        # Subtotal
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1, value=f"{cat_label} — Subtotal").font = Font(
            bold=True, color=C_SLATE_DARK[2:], size=9, name="Calibri"
        )
        ws.cell(row=row, column=1).fill = _fill(C_SLATE_MID)
        ws.cell(row=row, column=1).alignment = _align("right")
        sub_cell = ws.cell(row=row, column=6, value=cat_total)
        sub_cell.number_format = '₹#,##0.00'
        sub_cell.font = Font(bold=True, color=C_SLATE_DARK[2:], size=9, name="Calibri")
        sub_cell.fill = _fill(C_SLATE_MID)
        sub_cell.alignment = _align("right")
        ws.row_dimensions[row].height = 16
        row += 1

    # Contingency
    contingency = float(estimation.contingency_cost or grand_total * 0.05)
    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row=row, column=1, value="Contingency @ 5%").font = _font(bold=True, size=9)
    ws.cell(row=row, column=1).fill = _fill(C_SLATE_LIGHT)
    ws.cell(row=row, column=1).alignment = _align("right")
    cg_cell = ws.cell(row=row, column=6, value=contingency)
    cg_cell.number_format = '₹#,##0.00'
    cg_cell.font = _font(bold=True, size=9)
    cg_cell.fill = _fill(C_SLATE_LIGHT)
    cg_cell.alignment = _align("right")
    ws.row_dimensions[row].height = 16
    row += 1

    # Grand total
    total = float(estimation.total_cost or 0)
    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(
        bold=True, color=C_WHITE[2:], size=11, name="Calibri"
    )
    ws.cell(row=row, column=1).fill = _fill(C_DARK_BG)
    ws.cell(row=row, column=1).alignment = _align("right")
    gt_cell = ws.cell(row=row, column=6, value=total)
    gt_cell.number_format = '₹#,##0.00'
    gt_cell.font = Font(bold=True, color=C_BRAND_ORANGE[2:], size=12, name="Calibri")
    gt_cell.fill = _fill(C_DARK_BG)
    gt_cell.alignment = _align("right")
    ws.row_dimensions[row].height = 22

    # ── Column widths ──────────────────────────────────────────────────────────
    _set_col_width(ws, "A", 12)
    _set_col_width(ws, "B", 52)
    _set_col_width(ws, "C", 8)
    _set_col_width(ws, "D", 12)
    _set_col_width(ws, "E", 16)
    _set_col_width(ws, "F", 18)
    _set_col_width(ws, "G", 8)


def _build_category_sheet(wb: Workbook, estimation):
    ws = wb.create_sheet("📈 Category Totals")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    ws["A1"] = "Category Cost Totals"
    ws["A1"].fill = _fill(C_DARK_BG)
    ws["A1"].font = Font(bold=True, color=C_BRAND_ORANGE[2:], size=12, name="Calibri")
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 26

    for ci, h in enumerate(["Category","Amount (₹)","% Share"], start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill = _fill(C_SLATE_MID)
        cell.font = _font(bold=True)
        cell.alignment = _align("center")
        cell.border = _border()

    grouped: dict[str, float] = {}
    for item in estimation.boq_items:
        c = item.category.lower()
        grouped[c] = grouped.get(c, 0) + float(item.amount or 0)

    cats = sorted(grouped.keys(), key=lambda x: CAT_ORDER.index(x) if x in CAT_ORDER else 99)
    total = float(estimation.total_cost or 1)

    data_rows = []   # for chart

    row = 4
    for cat in cats:
        val = grouped[cat]
        pct = val/total*100
        label = CATEGORY_LABELS.get(cat, cat.title())
        data_rows.append((label, val))

        ws.cell(row=row, column=1, value=label).font = _font(size=10)
        ws.cell(row=row, column=1).fill = _fill(C_SLATE_LIGHT if row%2==0 else C_WHITE)
        ws.cell(row=row, column=1).border = _border()

        amt_cell = ws.cell(row=row, column=2, value=val)
        amt_cell.number_format = '₹#,##0.00'
        amt_cell.font = _font(size=10)
        amt_cell.alignment = _align("right")
        amt_cell.fill = _fill(C_SLATE_LIGHT if row%2==0 else C_WHITE)
        amt_cell.border = _border()

        pct_cell = ws.cell(row=row, column=3, value=round(pct,1))
        pct_cell.number_format = '0.0"%"'
        pct_cell.font = _font(size=10)
        pct_cell.alignment = _align("center")
        pct_cell.fill = _fill(C_SLATE_LIGHT if row%2==0 else C_WHITE)
        pct_cell.border = _border()
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, color=C_WHITE[2:], size=10, name="Calibri")
    ws.cell(row=row, column=1).fill = _fill(C_DARK_BG)
    ws.cell(row=row, column=1).border = _border()
    tot_c = ws.cell(row=row, column=2, value=total)
    tot_c.number_format = '₹#,##0.00'
    tot_c.font = Font(bold=True, color=C_BRAND_ORANGE[2:], size=10, name="Calibri")
    tot_c.alignment = _align("right")
    tot_c.fill = _fill(C_DARK_BG)
    tot_c.border = _border()
    ws.cell(row=row, column=3, value="100%").font = Font(bold=True, color=C_WHITE[2:], size=9, name="Calibri")
    ws.cell(row=row, column=3).fill = _fill(C_DARK_BG)
    ws.cell(row=row, column=3).alignment = _align("center")
    ws.cell(row=row, column=3).border = _border()

    _set_col_width(ws, "A", 22)
    _set_col_width(ws, "B", 20)
    _set_col_width(ws, "C", 12)

    # ── Bar chart ─────────────────────────────────────────────────────────────
    if data_rows:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Cost by Category"
        chart.y_axis.title = "Category"
        chart.x_axis.title = "Amount (₹)"
        chart.style = 10
        chart.width = 18
        chart.height = 12

        data_ref = Reference(ws, min_col=2, min_row=3, max_row=3+len(data_rows)-1)
        cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3+len(data_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        ws.add_chart(chart, f"E3")
